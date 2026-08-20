from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

CONTROL_SHEET_NAME = "Rename_Control"
LOG_SHEET_NAME = "Rename_Log"

INVALID_CHARS = r'<>:"/\\|?*'

# Windows reserved device names
WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9"
}


# ============================================================
# TKINTER SETUP
# ============================================================

def create_root():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    return root


# ============================================================
# FILE / FOLDER PICKERS
# ============================================================

def select_folder(root, title="Select Folder"):
    folder = filedialog.askdirectory(
        parent=root,
        title=title,
        mustexist=True
    )

    if not folder:
        return None

    return Path(folder)


def select_excel_file(root):
    file_path = filedialog.askopenfilename(
        parent=root,
        title="Select Excel Control File",
        filetypes=[
            ("Excel Files", "*.xlsx"),
            ("Excel Macro-Enabled Files", "*.xlsm"),
            ("All Files", "*.*")
        ]
    )

    if not file_path:
        return None

    return Path(file_path)


# ============================================================
# EXCEL FORMATTING
# ============================================================

def format_worksheet(ws):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 40,
        "B": 12,
        "C": 12,
        "D": 40,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def auto_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)

            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            60
        )


# ============================================================
# PHASE 1
# EXPORT FOLDER CONTENT TO EXCEL
# ============================================================

def extract_folder_to_excel(folder_path):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_file = folder_path / f"Rename_Control_{timestamp}.xlsx"

    items = []

    try:
        for item in folder_path.iterdir():

            # Only direct children
            if item.is_file():
                item_type = "File"

                # Preserve extension exactly as stored
                extension = item.suffix if item.suffix else "N/A"

            elif item.is_dir():
                item_type = "Folder"
                extension = "N/A"

            else:
                # Skip unusual filesystem objects
                continue

            items.append({
                "Current Name": item.name,
                "Type": item_type,
                "Extension": extension,
                "New Name": ""
            })

    except PermissionError:
        raise PermissionError(
            f"Permission denied while reading:\n{folder_path}"
        )

    # Sort for predictable output
    items.sort(
        key=lambda x: x["Current Name"].lower()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = CONTROL_SHEET_NAME

    headers = [
        "Current Name",
        "Type",
        "Extension",
        "New Name"
    ]

    ws.append(headers)

    for item in items:
        ws.append([
            item["Current Name"],
            item["Type"],
            item["Extension"],
            item["New Name"]
        ])

    format_worksheet(ws)

    # Add instructions sheet
    instruction_ws = wb.create_sheet("Instructions")

    instructions = [
        ["Bulk Rename Control File"],
        [""],
        ["Instructions"],
        ["1. Do not modify the Current Name column."],
        ["2. Enter the desired replacement name in New Name."],
        ["3. For files, you may omit the extension if you want to preserve the existing extension."],
        ["4. Example: Report.xlsx → Report_2026 becomes Report_2026.xlsx"],
        ["5. If you explicitly enter a different extension, it will be flagged for confirmation."],
        ["6. Changing an extension does NOT convert the file format."],
        ["7. Do not use Windows-invalid characters: < > : \" / \\ | ? *"],
        ["8. Save the workbook before running the rename operation."]
    ]

    for row in instructions:
        instruction_ws.append(row)

    instruction_ws.column_dimensions["A"].width = 100

    try:
        wb.save(output_file)
    except PermissionError:
        raise PermissionError(
            f"Could not create Excel file:\n{output_file}\n\n"
            "Check whether the file is already open."
        )

    return output_file, len(items)


# ============================================================
# VALIDATION HELPERS
# ============================================================

def is_blank(value):
    return value is None or str(value).strip() == ""


def validate_windows_name(name):
    """
    Validate a single Windows file/folder name.
    """

    if not name:
        return False, "New name is blank"

    name = str(name)

    # Invalid Windows characters
    for char in INVALID_CHARS:
        if char in name:
            return False, (
                f"Invalid Windows character '{char}'"
            )

    # Control characters
    for char in name:
        if ord(char) < 32:
            return False, "Name contains a control character"

    # Windows does not allow trailing spaces/dots
    if name.endswith(" ") or name.endswith("."):
        return False, (
            "Windows names cannot end with a space or period"
        )

    # Reserved Windows names
    base_name = name.split(".")[0].upper()

    if base_name in WINDOWS_RESERVED_NAMES:
        return False, (
            f"'{base_name}' is a reserved Windows name"
        )

    return True, ""


def get_extension(name):
    """
    Return the final extension.

    Examples:
        Report.xlsx -> .xlsx
        Report.tar.gz -> .gz
        Report       -> ""
        .gitignore   -> ""
    """

    return Path(name).suffix


def prepare_target_name(current_name, new_name, item_type):
    """
    Determine the actual target name.

    For files:
        Existing: ABC.xlsx
        New: Pump
        Result: Pump.xlsx

        Existing: ABC.xlsx
        New: Pump.pdf
        Result: Pump.pdf
        Extension change = True

    For folders:
        New name is used exactly as entered.
    """

    current_name = str(current_name).strip()
    new_name = str(new_name).strip()

    if item_type == "Folder":
        return new_name, False

    current_extension = get_extension(current_name)
    new_extension = get_extension(new_name)

    # No extension supplied:
    # preserve original extension.
    if new_extension == "":
        if current_extension:
            return new_name + current_extension, False

        return new_name, False

    # Explicit extension supplied.
    extension_changed = (
        current_extension.lower() != new_extension.lower()
    )

    return new_name, extension_changed


# ============================================================
# EXCEL CONTROL FILE READING
# ============================================================

def read_control_file(excel_path):
    try:
        wb = load_workbook(
            excel_path,
            read_only=True,
            data_only=True
        )
    except PermissionError:
        raise PermissionError(
            "The Excel control file could not be opened.\n"
            "Close it in Excel and try again."
        )
    except Exception as e:
        raise RuntimeError(
            f"Could not read Excel control file:\n{e}"
        )

    if CONTROL_SHEET_NAME not in wb.sheetnames:
        wb.close()

        raise ValueError(
            f"Required worksheet '{CONTROL_SHEET_NAME}' "
            f"was not found."
        )

    ws = wb[CONTROL_SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))

    wb.close()

    if not rows:
        raise ValueError("The control sheet is empty.")

    headers = [
        str(value).strip() if value is not None else ""
        for value in rows[0]
    ]

    required_columns = [
        "Current Name",
        "Type",
        "Extension",
        "New Name"
    ]

    missing_columns = [
        col for col in required_columns
        if col not in headers
    ]

    if missing_columns:
        raise ValueError(
            "Missing required Excel columns:\n"
            + "\n".join(missing_columns)
        )

    column_index = {
        header: headers.index(header)
        for header in required_columns
    }

    records = []

    for excel_row_number, row in enumerate(rows[1:], start=2):

        # Skip completely empty rows
        if all(
            value is None or str(value).strip() == ""
            for value in row
        ):
            continue

        def get_value(column_name):
            index = column_index[column_name]

            if index >= len(row):
                return ""

            value = row[index]

            if value is None:
                return ""

            return str(value).strip()

        records.append({
            "Excel Row": excel_row_number,
            "Current Name": get_value("Current Name"),
            "Type": get_value("Type"),
            "Extension": get_value("Extension"),
            "New Name": get_value("New Name")
        })

    return records


# ============================================================
# VALIDATION
# ============================================================

def validate_records(folder_path, records):
    results = []

    # Map exact current names to filesystem items.
    # Windows is case-insensitive, so use casefold for lookup,
    # while retaining exact displayed names.
    filesystem_items = {}

    try:
        for item in folder_path.iterdir():
            filesystem_items[item.name.casefold()] = item
    except PermissionError:
        raise PermissionError(
            f"Permission denied while reading:\n{folder_path}"
        )

    # Track requested target names
    requested_targets = {}

    for record in records:

        current_name = record["Current Name"]
        entered_new_name = record["New Name"]
        excel_row = record["Excel Row"]

        result = {
            "Excel Row": excel_row,
            "Current Name": current_name,
            "New Name": "",
            "Type": record["Type"],
            "Status": "",
            "Message": "",
            "Timestamp": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            "Source Path": "",
            "Target Path": "",
            "Extension Change": "No"
        }

        # ----------------------------------------------------
        # Blank current name
        # ----------------------------------------------------

        if not current_name:
            result["Status"] = "Failed"
            result["Message"] = "Current Name is blank"
            results.append(result)
            continue

        # ----------------------------------------------------
        # Blank new name
        # ----------------------------------------------------

        if is_blank(entered_new_name):
            result["Status"] = "Skipped"
            result["Message"] = "New name not provided"
            results.append(result)
            continue

        # ----------------------------------------------------
        # Locate exact current item
        # ----------------------------------------------------

        source_item = filesystem_items.get(
            current_name.casefold()
        )

        if source_item is None:
            result["Status"] = "Not Found"
            result["Message"] = "Current file/folder not found"
            results.append(result)
            continue

        result["Source Path"] = str(source_item)

        actual_type = (
            "Folder" if source_item.is_dir() else "File"
        )

        result["Type"] = actual_type

        # ----------------------------------------------------
        # Check type consistency
        # ----------------------------------------------------

        expected_type = record["Type"]

        if expected_type and expected_type != actual_type:
            result["Status"] = "Failed"
            result["Message"] = (
                f"Type mismatch: Excel says '{expected_type}', "
                f"but filesystem item is '{actual_type}'"
            )
            results.append(result)
            continue

        # ----------------------------------------------------
        # Prepare target name
        # ----------------------------------------------------

        target_name, extension_changed = prepare_target_name(
            current_name,
            entered_new_name,
            actual_type
        )

        result["New Name"] = target_name

        if extension_changed:
            result["Extension Change"] = "Yes"

        # ----------------------------------------------------
        # Validate Windows name
        # ----------------------------------------------------

        valid_name, validation_message = validate_windows_name(
            target_name
        )

        if not valid_name:
            result["Status"] = "Failed"
            result["Message"] = validation_message
            results.append(result)
            continue

        # ----------------------------------------------------
        # Same name
        # ----------------------------------------------------

        if current_name.casefold() == target_name.casefold():
            result["Status"] = "Skipped"
            result["Message"] = "New name is identical to current name"
            results.append(result)
            continue

        # ----------------------------------------------------
        # Check duplicate requested targets
        # ----------------------------------------------------

        target_key = target_name.casefold()

        if target_key in requested_targets:
            previous_row = requested_targets[target_key]

            result["Status"] = "Failed"
            result["Message"] = (
                f"Duplicate target name requested. "
                f"Same target is used by Excel row {previous_row}."
            )

            results.append(result)
            continue

        requested_targets[target_key] = excel_row

        # ----------------------------------------------------
        # Check if target already exists
        # ----------------------------------------------------

        target_path = folder_path / target_name

        if target_path.exists():
            result["Status"] = "Failed"
            result["Message"] = "Target already exists"
            results.append(result)
            continue

        result["Target Path"] = str(target_path)

        # ----------------------------------------------------
        # Extension warning
        # ----------------------------------------------------

        if extension_changed:
            result["Status"] = "Warning"
            result["Message"] = (
                "Extension change detected. "
                "This renames the extension only; "
                "it does NOT convert the file format."
            )
        else:
            result["Status"] = "Ready"
            result["Message"] = "Ready to rename"

        results.append(result)

    return results


# ============================================================
# DISPLAY VALIDATION SUMMARY
# ============================================================

def print_validation_summary(results):
    print("\n")
    print("=" * 90)
    print("VALIDATION SUMMARY")
    print("=" * 90)

    counters = {
        "Ready": 0,
        "Warning": 0,
        "Skipped": 0,
        "Failed": 0,
        "Not Found": 0
    }

    for result in results:
        status = result["Status"]

        if status in counters:
            counters[status] += 1

    for key, value in counters.items():
        print(f"{key:<15}: {value}")

    print("=" * 90)

    # Preview only executable rows
    preview_rows = [
        result for result in results
        if result["Status"] in ("Ready", "Warning")
    ]

    if preview_rows:
        print("\nRENAME PREVIEW")
        print("-" * 90)

        for index, result in enumerate(preview_rows, start=1):
            print(f"\n{index}. {result['Current Name']}")
            print(f"   → {result['New Name']}")

            if result["Extension Change"] == "Yes":
                print(
                    "   WARNING: Extension change detected"
                )

    print("\n")


# ============================================================
# CONFIRMATION
# ============================================================

def ask_confirmation(root, title, message):
    return messagebox.askyesno(
        title,
        message,
        parent=root
    )


# ============================================================
# PERFORM RENAME
# ============================================================

def execute_renames(results):
    execution_time = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    final_results = []

    for result in results:

        result = result.copy()

        if result["Status"] not in ("Ready", "Warning"):
            final_results.append(result)
            continue

        source_path = Path(result["Source Path"])
        target_path = Path(result["Target Path"])

        try:

            # Re-check before rename.
            if not source_path.exists():
                result["Status"] = "Failed"
                result["Message"] = (
                    "Source disappeared before rename"
                )

                final_results.append(result)
                continue

            # Never overwrite.
            if target_path.exists():
                result["Status"] = "Failed"
                result["Message"] = (
                    "Target exists at execution time"
                )

                final_results.append(result)
                continue

            source_path.rename(target_path)

            result["Status"] = "Renamed"
            result["Message"] = "Rename successful"
            result["Timestamp"] = execution_time

        except PermissionError:
            result["Status"] = "Failed"
            result["Message"] = (
                "Permission denied. File/folder may be "
                "protected or in use."
            )

        except FileNotFoundError:
            result["Status"] = "Not Found"
            result["Message"] = (
                "Source file/folder no longer exists"
            )

        except FileExistsError:
            result["Status"] = "Failed"
            result["Message"] = (
                "Target already exists"
            )

        except OSError as e:
            result["Status"] = "Failed"
            result["Message"] = (
                f"Windows filesystem error: {e}"
            )

        except Exception as e:
            result["Status"] = "Failed"
            result["Message"] = (
                f"Unexpected error: {e}"
            )

        final_results.append(result)

    return final_results


# ============================================================
# CREATE LOG EXCEL
# ============================================================

def create_log_file(folder_path, results):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    log_path = folder_path / f"Rename_Log_{timestamp}.xlsx"

    wb = Workbook()

    ws = wb.active
    ws.title = LOG_SHEET_NAME

    headers = [
        "Excel Row",
        "Current Name",
        "New Name",
        "Type",
        "Status",
        "Message",
        "Extension Change",
        "Timestamp",
        "Source Path",
        "Target Path"
    ]

    ws.append(headers)

    for result in results:
        ws.append([
            result.get("Excel Row", ""),
            result.get("Current Name", ""),
            result.get("New Name", ""),
            result.get("Type", ""),
            result.get("Status", ""),
            result.get("Message", ""),
            result.get("Extension Change", ""),
            result.get("Timestamp", ""),
            result.get("Source Path", ""),
            result.get("Target Path", "")
        ])

    format_worksheet(ws)
    auto_width(ws)

    # Summary sheet
    summary = wb.create_sheet("Summary")

    summary.append(["Rename Automation Summary"])
    summary.append([""])
    summary.append([
        "Generated",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])
    summary.append(["Folder", str(folder_path)])
    summary.append([""])

    statuses = [
        "Renamed",
        "Skipped",
        "Failed",
        "Warning",
        "Not Found",
        "Ready"
    ]

    summary.append(["Status", "Count"])

    for status in statuses:
        count = sum(
            1
            for result in results
            if result["Status"] == status
        )

        summary.append([status, count])

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 80

    try:
        wb.save(log_path)
    except PermissionError:
        raise PermissionError(
            f"Could not save log file:\n{log_path}\n\n"
            "Make sure the log file is not open in Excel."
        )

    return log_path


# ============================================================
# PHASE 2
# RENAME WORKFLOW
# ============================================================

def rename_workflow(root):
    print("\n")
    print("=" * 70)
    print("BULK FILE / FOLDER RENAME")
    print("=" * 70)

    # --------------------------------------------------------
    # Select folder
    # --------------------------------------------------------

    folder_path = select_folder(
        root,
        "Select Folder Containing Files/Folders to Rename"
    )

    if not folder_path:
        print("Operation cancelled.")
        return

    print(f"\nSelected folder:\n{folder_path}")

    # --------------------------------------------------------
    # Select control Excel
    # --------------------------------------------------------

    excel_path = select_excel_file(root)

    if not excel_path:
        print("Operation cancelled.")
        return

    print(f"\nControl file:\n{excel_path}")

    # --------------------------------------------------------
    # Read Excel
    # --------------------------------------------------------

    try:
        records = read_control_file(excel_path)
    except Exception as e:
        messagebox.showerror(
            "Control File Error",
            str(e),
            parent=root
        )
        return

    print(
        f"\nLoaded {len(records)} control rows."
    )

    # --------------------------------------------------------
    # Validate
    # --------------------------------------------------------

    try:
        results = validate_records(
            folder_path,
            records
        )
    except Exception as e:
        messagebox.showerror(
            "Validation Error",
            str(e),
            parent=root
        )
        return

    # --------------------------------------------------------
    # Display summary
    # --------------------------------------------------------

    print_validation_summary(results)

    ready_count = sum(
        1
        for result in results
        if result["Status"] in ("Ready", "Warning")
    )

    failed_count = sum(
        1
        for result in results
        if result["Status"] in ("Failed", "Not Found")
    )

    warning_count = sum(
        1
        for result in results
        if result["Status"] == "Warning"
    )

    # --------------------------------------------------------
    # Safety rule:
    # Any validation failure blocks execution.
    # --------------------------------------------------------

    if failed_count > 0:

        messagebox.showerror(
            "Validation Failed",
            f"{failed_count} row(s) contain validation errors.\n\n"
            "No files or folders were renamed.\n\n"
            "Please correct the Excel control file and run "
            "the automation again.",
            parent=root
        )

        # Still create a validation log
        try:
            log_path = create_log_file(
                folder_path,
                results
            )

            print(
                f"\nValidation report created:\n{log_path}"
            )

        except Exception as e:
            print(
                f"\nCould not create validation report:\n{e}"
            )

        return

    # --------------------------------------------------------
    # Nothing to rename
    # --------------------------------------------------------

    if ready_count == 0:

        messagebox.showinfo(
            "Nothing to Rename",
            "There are no valid rename operations.",
            parent=root
        )

        return

    # --------------------------------------------------------
    # Extension warning
    # --------------------------------------------------------

    if warning_count > 0:

        extension_message = (
            f"{warning_count} file(s) have an extension change.\n\n"
            "IMPORTANT:\n\n"
            "Changing a filename extension does NOT convert "
            "the file into another format.\n\n"
            "Example:\n"
            "Report.xlsx → Report.pdf\n\n"
            "This only changes the filename. The contents "
            "remain an Excel file.\n\n"
            "Do you explicitly want to allow these "
            "extension changes?"
        )

        allow_extension_change = ask_confirmation(
            root,
            "Extension Change Warning",
            extension_message
        )

        if not allow_extension_change:

            messagebox.showinfo(
                "Cancelled",
                "Rename operation cancelled because "
                "extension changes were not approved.",
                parent=root
            )

            return

    # --------------------------------------------------------
    # Final confirmation
    # --------------------------------------------------------

    confirmation_message = (
        f"Ready to rename {ready_count} item(s).\n\n"
        f"Folder:\n{folder_path}\n\n"
        "No existing files/folders will be overwritten.\n\n"
        "Do you want to proceed?"
    )

    proceed = ask_confirmation(
        root,
        "Confirm Rename",
        confirmation_message
    )

    if not proceed:

        print("Rename cancelled by user.")

        messagebox.showinfo(
            "Cancelled",
            "No files or folders were renamed.",
            parent=root
        )

        return

    # --------------------------------------------------------
    # Execute
    # --------------------------------------------------------

    print("\nExecuting rename operations...")

    final_results = execute_renames(results)

    # --------------------------------------------------------
    # Create final log
    # --------------------------------------------------------

    try:
        log_path = create_log_file(
            folder_path,
            final_results
        )
    except Exception as e:
        messagebox.showerror(
            "Log Error",
            f"Renaming completed, but the log could not "
            f"be created.\n\n{e}",
            parent=root
        )

        return

    # --------------------------------------------------------
    # Final summary
    # --------------------------------------------------------

    renamed = sum(
        1
        for result in final_results
        if result["Status"] == "Renamed"
    )

    failed = sum(
        1
        for result in final_results
        if result["Status"] == "Failed"
    )

    skipped = sum(
        1
        for result in final_results
        if result["Status"] == "Skipped"
    )

    not_found = sum(
        1
        for result in final_results
        if result["Status"] == "Not Found"
    )

    final_message = (
        "Rename operation completed.\n\n"
        f"Renamed: {renamed}\n"
        f"Failed: {failed}\n"
        f"Skipped: {skipped}\n"
        f"Not Found: {not_found}\n\n"
        f"Log:\n{log_path}"
    )

    print("\n" + "=" * 70)
    print("RENAME COMPLETE")
    print("=" * 70)
    print(final_message)

    messagebox.showinfo(
        "Rename Complete",
        final_message,
        parent=root
    )


# ============================================================
# MAIN MENU
# ============================================================

def main():
    root = create_root()

    print("\n")
    print("=" * 70)
    print("WINDOWS BULK FILE / FOLDER RENAMER")
    print("=" * 70)

    print("\nChoose an operation:")
    print()
    print("1. Extract folder names to Excel")
    print("2. Rename using Excel control file")
    print("3. Exit")
    print()

    while True:

        choice = input(
            "Enter your choice [1/2/3]: "
        ).strip()

        if choice == "1":

            folder_path = select_folder(
                root,
                "Select Folder to Export"
            )

            if not folder_path:
                print("Operation cancelled.")
                break

            try:
                output_file, count = (
                    extract_folder_to_excel(folder_path)
                )

                message = (
                    f"Excel control file created successfully.\n\n"
                    f"Items extracted: {count}\n\n"
                    f"File:\n{output_file}"
                )

                print("\n" + message)

                messagebox.showinfo(
                    "Export Complete",
                    message,
                    parent=root
                )

            except Exception as e:

                print(f"\nError: {e}")

                messagebox.showerror(
                    "Export Error",
                    str(e),
                    parent=root
                )

            break

        elif choice == "2":

            rename_workflow(root)
            break

        elif choice == "3":

            print("Goodbye.")
            break

        else:
            print(
                "Invalid choice. Please enter 1, 2 or 3."
            )

    root.destroy()


if __name__ == "__main__":
    main()